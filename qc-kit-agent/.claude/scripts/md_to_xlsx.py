#!/usr/bin/env python3
"""
Convert markdown artifact to Excel.
- Generic artifacts (analysis.md, modules.md, ...): each ## section -> 1 sheet,
  built from scratch (bold header, freeze row, wrap text, <br> -> newline)
- test-cases.md: loaded on top of the real company template (Web/App), keeping
  existing sheets/dropdowns.
  - Mỗi ## section có bảng TC (cột ID + Test Scenario) -> 1 sheet TC riêng
    (clone từ template, giữ dropdown), liệt kê vào Summary.
  - Section khác (Test Data Reference, ghi chú...) -> sheet generic, chèn
    cuối workbook, không đưa vào Summary.
"""

import sys
import re
import copy
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATE_DIR = Path(__file__).resolve().parent.parent.parent / "template"
TEMPLATE_FILES = {
    "web": "[Tên dự án]_[Tên module]_Testcase_ForWeb_V3.0.xlsx",
    "app": "[Tên dự án]_[Tên module]_Testcase_ForApp_V3.0.xlsx",
}
TC_SHEET_NAME = "Screen name"
TC_EVIDENCE_SHEET_NAME = "Screen name_evidence"
SUMMARY_SHEET_NAME = "Summary"
SUMMARY_SCREEN_NAME_COL = 3  # cột C — formula "TCs plan" đọc tên sheet TC từ đây
SUMMARY_FIRST_ROW = 4
SUMMARY_LAST_ROW = 17  # template chỉ dựng sẵn formula Summary tới dòng này (fixed range)
# md header (lowercase) -> template header (lowercase) alias, khi tên không khớp 1-1
TC_HEADER_ALIASES = {
    "risk level": "risk",
}
# Formula/dropdown trong template chỉ có sẵn tới dòng này (fixed range, đã confirm với user)
TC_TEMPLATE_MAX_ROW = 97

HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="46BDC6", end_color="46BDC6", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
TEXT_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ORDINAL_PREFIX_RE = re.compile(r"^\d+\.\s*")


def parse_markdown(content):
    sections = []
    current = {"heading": None, "lines": []}

    for line in content.split("\n"):
        m = re.match(r"^#{1,3}\s+(.+)$", line)
        if m:
            if current["heading"] or current["lines"]:
                sections.append(current)
            current = {"heading": m.group(1).strip(), "lines": []}
        else:
            current["lines"].append(line)

    if current["heading"] or current["lines"]:
        sections.append(current)

    return sections


def is_table_row(line):
    return line.strip().startswith("|") and line.strip().endswith("|")


def is_separator_row(line):
    return bool(re.match(r"^\|[-:\s|]+\|$", line.strip()))


def parse_table(table_lines):
    headers = []
    rows = []
    for line in table_lines:
        if is_separator_row(line):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if not headers:
            headers = cells
        else:
            rows.append(cells)
    return headers, rows


def extract_blocks(lines):
    blocks = []
    i = 0
    text_buffer = []

    while i < len(lines):
        line = lines[i]
        if is_table_row(line):
            if text_buffer:
                text = "\n".join(text_buffer).strip()
                if text:
                    blocks.append({"type": "text", "content": text})
                text_buffer = []
            table_lines = []
            while i < len(lines) and is_table_row(lines[i]):
                table_lines.append(lines[i])
                i += 1
            headers, rows = parse_table(table_lines)
            if headers:
                blocks.append({"type": "table", "headers": headers, "rows": rows})
        else:
            text_buffer.append(line)
            i += 1

    if text_buffer:
        text = "\n".join(text_buffer).strip()
        if text:
            blocks.append({"type": "text", "content": text})

    return blocks


def convert_br(text):
    return re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)


def clean_heading(heading):
    """Bỏ số thứ tự đầu heading (VD "1. AUTH-01 — ..." -> "AUTH-01 — ...")
    để đỡ tốn ký tự trong giới hạn 31 ký tự tên sheet của Excel."""
    return ORDINAL_PREFIX_RE.sub("", (heading or "").strip())


SHEET_NAME_LIMIT = 31
CODE_DESC_RE = re.compile(r"^(.{1,24}?)\s+—\s+(.+)$")


def strip_forbidden_chars(name):
    # Excel cấm các ký tự này trong tên sheet — thay bằng space (không xoá hẳn)
    # để tránh 2 từ dính liền nhau (VD "Logout / Đổi" -> "Logout   Đổi", không phải "LogoutĐổi")
    name = re.sub(r"[\\/*?:\[\]]", " ", name)
    return re.sub(r"\s+", " ", name).strip()


def truncate_at_word(text, limit):
    text = text.strip()
    if len(text) <= limit:
        return text
    cut = text[:limit]
    last_space = cut.rfind(" ")
    if last_space > 0:
        cut = cut[:last_space]
    return cut.rstrip(" ([{,;.:—-")


def smart_sheet_name(heading, limit=SHEET_NAME_LIMIT):
    """Rút gọn heading thành tên sheet hợp lệ (<=31 ký tự), ưu tiên giữ nguyên phần
    "mã" ngắn gọn đứng trước dấu " — " (VD "AUTH-07"), cắt phần mô tả phía sau tại
    ranh giới từ (không cắt giữa từ) để phần còn lại vẫn đọc được, thay vì cắt cứng
    theo số ký tự."""
    heading = strip_forbidden_chars(heading)
    if not heading:
        return "Sheet"
    if len(heading) <= limit:
        return heading

    m = CODE_DESC_RE.match(heading)
    if m:
        code, desc = m.group(1), m.group(2)
        sep = " — "
        budget = limit - len(code) - len(sep)
        if budget > 0:
            desc_trunc = truncate_at_word(desc, budget)
            result = f"{code}{sep}{desc_trunc}".rstrip(" —") if desc_trunc else code
            return result[:limit] or "Sheet"
        return code[:limit] or "Sheet"

    return truncate_at_word(heading, limit) or heading[:limit] or "Sheet"


def unique_sheet_name(wb, name):
    base = name
    count = 1
    existing = [s.title for s in wb.worksheets]
    while name in existing:
        name = f"{base[:28]}_{count}"
        count += 1
    return name


def is_test_cases_file(input_path):
    return input_path.stem.strip().lower() == "test-cases"


def guess_module_name(content, input_path):
    m = re.search(r"^#\s+.*?:\s*(.+)$", content, re.MULTILINE)
    if m:
        return m.group(1).strip()
    parent = input_path.resolve().parent.name
    return parent if parent and parent.lower() != "testing" else input_path.stem


def is_tc_table(headers):
    lower_headers = [h.strip().lower() for h in headers]
    return "test scenario" in lower_headers and "id" in lower_headers


def classify_sections(content, fallback_heading):
    """Tách sections thành 2 nhóm:
    - tc_sections: có bảng TC (cột ID + Test Scenario) -> {"heading", "headers", "rows"}
    - other_sections: còn lại (free text / bảng khác, VD Test Data Reference) -> {"heading", "lines"}
    """
    sections = parse_markdown(content)
    tc_sections = []
    other_sections = []

    for section in sections:
        heading = section["heading"] or fallback_heading
        blocks = extract_blocks(section["lines"])
        tc_headers = None
        tc_rows = []
        for block in blocks:
            if block["type"] == "table" and is_tc_table(block["headers"]):
                tc_headers = block["headers"]
                tc_rows.extend(block["rows"])

        if tc_headers:
            tc_sections.append({"heading": heading, "headers": tc_headers, "rows": tc_rows})
        elif blocks:
            other_sections.append({"heading": heading, "lines": section["lines"]})

    return tc_sections, other_sections


def build_col_map(headers):
    """Map index cột trong bảng md (theo tên header) -> index cột trong template TC sheet."""
    col_map = {}
    unmatched = []
    for i, h in enumerate(headers):
        key = h.strip().lower()
        key = TC_HEADER_ALIASES.get(key, key)
        col_map[i] = key
    return col_map


def populate_generic_sheet(ws, lines):
    ws.sheet_view.showGridLines = True
    row = 1

    for block in extract_blocks(lines):
        if block["type"] == "text":
            cell = ws.cell(row=row, column=1, value=block["content"])
            cell.alignment = TEXT_ALIGN
            cell.font = Font(size=10, italic=True, color="555555")
            row += 2

        elif block["type"] == "table":
            headers = block["headers"]
            rows = block["rows"]

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=convert_br(h))
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGN
                cell.border = BORDER

            ws.freeze_panes = ws.cell(row=row + 1, column=1)
            ws.row_dimensions[row].height = 24
            row += 1

            for r_data in rows:
                for col, val in enumerate(r_data, 1):
                    cell = ws.cell(row=row, column=col, value=convert_br(val))
                    cell.alignment = CELL_ALIGN
                    cell.border = BORDER
                row += 1

            row += 1  # blank row after table

    for col_cells in ws.columns:
        max_w = 10
        for cell in col_cells:
            if cell.value:
                for part in str(cell.value).split("\n"):
                    max_w = max(max_w, len(part) * 1.1)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max_w, 55)


def write_xlsx(sections, output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for section in sections:
        sheet_name = unique_sheet_name(wb, smart_sheet_name(clean_heading(section["heading"]) or "Sheet"))
        ws = wb.create_sheet(title=sheet_name)
        populate_generic_sheet(ws, section["lines"])

    wb.save(output_path)


def write_tc_from_template(template_path, tc_sections, other_sections, output_path):
    wb = openpyxl.load_workbook(template_path)

    template_ws = wb[TC_SHEET_NAME]
    template_evidence_ws = wb[TC_EVIDENCE_SHEET_NAME] if TC_EVIDENCE_SHEET_NAME in wb.sheetnames else None
    base_dvs = list(template_ws.data_validations.dataValidation)

    header_to_col = {}
    for c in range(1, template_ws.max_column + 1):
        v = template_ws.cell(row=1, column=c).value
        if v:
            header_to_col[str(v).strip().lower()] = c

    unmatched = set()
    tc_sheet_titles = []
    ordered_sheet_names = []  # thứ tự sheet mong muốn cuối cùng

    # Giữ nguyên 2 sheet đầu của template (Test view point, Summary) đúng vị trí gốc
    for name in wb.sheetnames:
        if name not in (TC_SHEET_NAME, TC_EVIDENCE_SHEET_NAME):
            ordered_sheet_names.append(name)

    # Pha 1: tạo/clone TẤT CẢ sheet trước, trong khi template_ws còn pristine (chưa ghi data).
    # Nếu ghi data ngay trong vòng lặp rồi mới clone tiếp, sheet sau sẽ dính data thừa của
    # sheet trước — vì copy_worksheet() clone theo state HIỆN TẠI của nguồn.
    sheet_infos = []  # (title, ws, tc_section)
    for i, tc_section in enumerate(tc_sections):
        title = unique_sheet_name(wb, smart_sheet_name(clean_heading(tc_section["heading"])))
        if i == 0:
            ws = template_ws
            ws.title = title
        else:
            ws = wb.copy_worksheet(template_ws)
            ws.title = title
            for dv in base_dvs:
                ws.add_data_validation(copy.copy(dv))
        sheet_infos.append((title, ws, tc_section))

        tc_sheet_titles.append(title)
        ordered_sheet_names.append(title)

    # Chỉ giữ 1 sheet evidence duy nhất (dùng chung cho mọi sheet TC) — đặt ngay sau
    # sheet TC cuối cùng, không nhân bản theo từng sheet TC.
    if template_evidence_ws is not None:
        ordered_sheet_names.append(template_evidence_ws.title)

    # Pha 2: điền data — an toàn vì mọi sheet đã được clone xong từ bản pristine ở Pha 1.
    for title, ws, tc_section in sheet_infos:
        col_map = build_col_map(tc_section["headers"])
        real_col_map = {}
        for idx, key in col_map.items():
            if key in header_to_col:
                real_col_map[idx] = header_to_col[key]
            else:
                unmatched.add(tc_section["headers"][idx])

        for r_i, row_vals in enumerate(tc_section["rows"]):
            r = 2 + r_i
            for idx, val in enumerate(row_vals):
                if idx in real_col_map:
                    ws.cell(row=r, column=real_col_map[idx], value=convert_br(val))
            if r > TC_TEMPLATE_MAX_ROW:
                # Ngoài phạm vi dropdown/formula cố định của template -> vẫn ghi số thứ tự thô
                ws.cell(row=r, column=1, value=r - 1)

    overflow = []
    if SUMMARY_SHEET_NAME in wb.sheetnames:
        summary_ws = wb[SUMMARY_SHEET_NAME]
        for i, title in enumerate(tc_sheet_titles):
            row = SUMMARY_FIRST_ROW + i
            if row > SUMMARY_LAST_ROW:
                overflow.append(title)
                continue
            # Chỉ ghi tên sheet vào cột "Screen name" — formula "TCs plan" (COUNTA/INDIRECT)
            # có sẵn trong template sẽ tự đếm lại, kể cả khi QC sửa TC thủ công sau này.
            cell = summary_ws.cell(row=row, column=SUMMARY_SCREEN_NAME_COL, value=title)
            # Riêng C4 có kèm hyperlink nội bộ trỏ sang sheet TC để member click nhảy nhanh —
            # phải cập nhật lại target/display, nếu không Google Sheets sẽ hiện lại text cache cũ.
            if cell.hyperlink is not None:
                cell.hyperlink.location = f"'{title}'!A1"
                cell.hyperlink.display = title

    # Sections không phải TC (Test Data Reference, ghi chú...) -> sheet generic, chèn cuối cùng
    for section in other_sections:
        title = unique_sheet_name(wb, smart_sheet_name(clean_heading(section["heading"]) or "Sheet"))
        ws = wb.create_sheet(title=title)
        populate_generic_sheet(ws, section["lines"])
        ordered_sheet_names.append(title)

    wb._sheets = [wb[name] for name in ordered_sheet_names]

    wb.save(output_path)
    return sorted(unmatched), overflow


def main():
    if len(sys.argv) < 2:
        print("Usage: md_to_xlsx.py <input.md> [web|app]")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    if not input_path.exists():
        print(f"File not found: {input_path}")
        sys.exit(1)

    platform = sys.argv[2].strip().lower() if len(sys.argv) > 2 else None
    output_path = input_path.with_suffix(".xlsx")
    content = input_path.read_text(encoding="utf-8")

    if is_test_cases_file(input_path):
        if platform not in TEMPLATE_FILES:
            print("NEED_PLATFORM: file này là test-cases.md, cần chỉ định platform 'web' hoặc 'app'.")
            sys.exit(2)

        template_path = TEMPLATE_DIR / TEMPLATE_FILES[platform]
        if not template_path.exists():
            print(f"Template not found: {template_path}")
            sys.exit(1)

        fallback_heading = guess_module_name(content, input_path)
        tc_sections, other_sections = classify_sections(content, fallback_heading)
        if not tc_sections:
            print("Không tìm thấy bảng Test Case hợp lệ (cần cột 'ID' và 'Test Scenario') trong file.")
            sys.exit(1)

        unmatched, overflow = write_tc_from_template(template_path, tc_sections, other_sections, output_path)
        total_tcs = sum(len(s["rows"]) for s in tc_sections)
        sheet_names = [clean_heading(s["heading"]) for s in tc_sections]
        print(f"Done ({platform} template, {total_tcs} TCs -> {len(tc_sections)} sheet: {sheet_names}): {output_path}")
        if other_sections:
            print(f"Đã thêm {len(other_sections)} sheet tham khảo (không thuộc Summary): "
                  f"{[clean_heading(s['heading']) for s in other_sections]}")
        if unmatched:
            print(f"Lưu ý: {len(unmatched)} cột không khớp template, đã bỏ qua: {unmatched}")
        if overflow:
            print(f"Lưu ý: {len(overflow)} sheet vượt quá {SUMMARY_LAST_ROW - SUMMARY_FIRST_ROW + 1} dòng "
                  f"Summary hỗ trợ, không được liệt kê: {overflow}")
    else:
        sections = parse_markdown(content)
        write_xlsx(sections, output_path)
        print(f"Done: {output_path}")


if __name__ == "__main__":
    main()
